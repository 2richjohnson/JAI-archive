#!/usr/bin/env python3
"""
08_export_excel.py — Reconstruct Excel workbooks from cost study facts in DuckDB.

For each cost study (grouped by _source_doc + study description), creates one
Excel worksheet with:
  - Line items as rows (row_label, value_raw, value_numeric, unit)
  - Simple SUM formulas inferred for subtotals/totals where structure is clear
  - Provenance note: source document, currency year, confidence

Usage:
    python 08_export_excel.py [--output DIR] [--source DOC]

    --output DIR    Output directory (default: ~/jai-archive/exports/excel/)
    --source DOC    Export only this source document (partial match)
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import duckdb

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path.home() / "jai-archive"
DUCKDB_PATH = BASE_DIR / "duckdb" / "jai.db"
DEFAULT_OUTPUT = BASE_DIR / "exports" / "excel"

# Style constants
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT = Font(bold=True, size=11)
SUBTOTAL_FILL = PatternFill("solid", fgColor="EBF3FB")
THIN = Side(border_style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cell(cell, fill=None, font=None, align="left", number_format=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = Alignment(horizontal=align)
    if number_format:
        cell.number_format = number_format
    cell.border = THIN_BORDER


def build_workbook(study_rows: list[dict], study_meta: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Study"

    # ── Header block ──────────────────────────────────────────────────────────
    ws["A1"] = "JAI Archive — Reconstructed Cost Study"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Source document: {study_meta.get('source_doc', 'unknown')}"
    ws["A3"] = f"Study: {study_meta.get('study', '')}"
    ws["A4"] = f"Study year: {study_meta.get('study_year', 'unknown')}   " \
               f"Currency year: {study_meta.get('currency_year', 'unknown')}"
    ws["A5"] = f"Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A6"] = (
        "NOTE: This reconstruction is derived from OCR'd PDF data. "
        "Formulas are inferred from line_item_type context — verify before financial use."
    )
    ws["A6"].font = Font(italic=True, color="AA0000")

    ws.row_dimensions[1].height = 20
    for row in range(1, 7):
        ws.merge_cells(f"A{row}:E{row}")

    # ── Column headers ────────────────────────────────────────────────────────
    header_row = 8
    headers = ["Line Item", "Value (raw)", "Value (numeric)", "Unit", "Type"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        style_cell(cell, fill=HDR_FILL, font=HDR_FONT, align="center")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Track row numbers for formula inference
    input_rows = []   # (excel_row, value_numeric) for input lines
    subtotal_start_rows = []

    data_start = header_row + 1
    current_excel_row = data_start

    for r in study_rows:
        row_label = r.get("row_label") or r.get("attribute", "")
        value_raw = r.get("value_raw", "")
        value_numeric = r.get("value_numeric")
        unit = r.get("unit", "")
        line_type = r.get("line_item_type") or ""

        is_total = line_type == "total"
        is_subtotal = line_type == "subtotal"
        is_calculated = line_type == "calculated"

        # Cell A: line item label
        cell_a = ws.cell(row=current_excel_row, column=1, value=row_label)
        # Cell B: raw value
        cell_b = ws.cell(row=current_excel_row, column=2, value=value_raw)
        # Cell C: numeric or inferred formula
        if is_total and input_rows:
            # Try to write a SUM formula over preceding input rows
            refs = [f"C{er}" for er, _ in input_rows]
            if len(refs) <= 20:
                formula = f"=SUM({','.join(refs)})"
            else:
                formula = f"=SUM(C{input_rows[0][0]}:C{input_rows[-1][0]})"
            cell_c = ws.cell(row=current_excel_row, column=3, value=formula)
        elif is_subtotal and subtotal_start_rows:
            start = subtotal_start_rows[-1]
            cell_c = ws.cell(row=current_excel_row, column=3,
                             value=f"=SUM(C{start}:C{current_excel_row - 1})")
        else:
            cell_c = ws.cell(row=current_excel_row, column=3, value=value_numeric)
        # Cell D: unit
        cell_d = ws.cell(row=current_excel_row, column=4, value=unit)
        # Cell E: type
        cell_e = ws.cell(row=current_excel_row, column=5, value=line_type)

        if is_total:
            fill, font = TOTAL_FILL, TOTAL_FONT
        elif is_subtotal:
            fill, font = SUBTOTAL_FILL, Font(bold=True)
        else:
            fill, font = None, None

        for cell in (cell_a, cell_b, cell_c, cell_d, cell_e):
            style_cell(cell, fill=fill, font=font)
        if value_numeric is not None:
            style_cell(cell_c, fill=fill, font=font, align="right", number_format="#,##0.0")

        # Track rows for formula inference
        if not is_total and not is_subtotal and not is_calculated and value_numeric is not None:
            input_rows.append((current_excel_row, value_numeric))
        if is_subtotal:
            subtotal_start_rows.append(current_excel_row + 1)

        current_excel_row += 1

    # ── Provenance sheet ──────────────────────────────────────────────────────
    ps = wb.create_sheet("Provenance")
    ps["A1"] = "Field"
    ps["B1"] = "Value"
    ps["A1"].font = Font(bold=True)
    ps["B1"].font = Font(bold=True)

    provenance = [
        ("Source document", study_meta.get("source_doc", "")),
        ("Study description", study_meta.get("study", "")),
        ("Study year", study_meta.get("study_year", "")),
        ("Currency year", study_meta.get("currency_year", "")),
        ("Confidence", study_meta.get("confidence", "")),
        ("Rows extracted", len(study_rows)),
        ("Export timestamp", datetime.utcnow().isoformat()),
        ("Warning", "Formulas are inferred approximations. Verify before financial use."),
        ("Formula coverage", "SUM formulas written for totals/subtotals where line_item_type is known."),
    ]
    for i, (k, v) in enumerate(provenance, 2):
        ps.cell(row=i, column=1, value=k)
        ps.cell(row=i, column=2, value=str(v))
    ps.column_dimensions["A"].width = 25
    ps.column_dimensions["B"].width = 60

    return wb


def export_studies(output_dir: Path, source_filter: str | None):
    if not DUCKDB_PATH.exists():
        print(f"DuckDB not found at {DUCKDB_PATH} — run 06_setup_duckdb.py first")
        sys.exit(1)

    con = duckdb.connect(str(DUCKDB_PATH), read_only=True)

    filter_clause = ""
    if source_filter:
        filter_clause = f"AND source_doc ILIKE '%{source_filter}%'"

    studies = con.execute(f"""
        SELECT DISTINCT source_doc, study, study_year, currency_year,
               COUNT(*) AS row_count, ANY_VALUE(_confidence) AS confidence
        FROM cost_summary
        WHERE 1=1 {filter_clause}
        GROUP BY source_doc, study, study_year, currency_year
        ORDER BY source_doc, study
    """).fetchall()

    if not studies:
        print("No cost studies found in DuckDB. Run extraction first.")
        con.close()
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Found {len(studies)} cost study/studies")

    for source_doc, study, study_year, currency_year, row_count, confidence in studies:
        rows = con.execute("""
            SELECT row_label, attribute, value_raw, value_numeric, unit, line_item_type
            FROM cost_summary
            WHERE source_doc = ? AND (study = ? OR (study IS NULL AND ? IS NULL))
            ORDER BY rowid
        """, [source_doc, study, study]).fetchdf().to_dict(orient="records")

        meta = {
            "source_doc": source_doc,
            "study": study,
            "study_year": study_year,
            "currency_year": currency_year,
            "confidence": confidence,
        }

        safe_name = re.sub(r"[^\w\-]", "_", f"{Path(source_doc).stem}_{study or 'cost'}")[:60]
        out_path = output_dir / f"{safe_name}.xlsx"

        wb = build_workbook(rows, meta)
        wb.save(out_path)
        print(f"  Saved: {out_path.name}  ({row_count} rows)")

    con.close()
    print(f"\nExports written to: {output_dir}")


def main():
    parser = argparse.ArgumentParser(description="Export JAI cost studies to Excel.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output directory")
    parser.add_argument("--source", help="Filter by source document name (partial match)")
    args = parser.parse_args()

    export_studies(Path(args.output), args.source)


if __name__ == "__main__":
    main()
